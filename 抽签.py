# -*- coding: utf-8 -*-
"""
Created on Fri Nov  6 12:39:01 2020

@author: 羽
"""

import random
import openpyxl
wb = openpyxl.load_workbook(r"./1_再来亿杯第一赛季积分.xlsx")
sheet = wb.get_sheet_by_name('Sheet1')
interator = 0
dic={}
data = []
for i in range(1,sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value is not None and isinstance(sheet.cell(row=i, column=1).value, int):
        dic.setdefault(sheet.cell(i,2).value,int(sheet.cell(i,1).value))
        if sheet.cell(i,4).value is None:
            sheet.cell(i,4).value = 1
print(dic)
wb.save(r"./1_再来亿杯第一赛季积分.xlsx")
# dic = {"郑雨楠": 1 ,"胡博兮": 2 ,"夏逸雨": 3 ,"汪锴": 4 ,"赵皓然": 5 ,"杨帅聪": 6 ,"刘孟乔": 7 ,"周书瑗": 8 ,"李寿媛": 9 ,"陈志辉": 10 ,"杨飒": 11 ,"黄凯威": 12 ,"蓝庭旭": 13 ,"杨致远": 14 ,"黄博闻": 15 ,"鲍云天": 16 ,"刘佳祺": 17 ,"陈千奕": 18 ,"周陇": 19 ,"厍妍": 20 ,"房任博": 21 ,"谭初石": 22 ,"黄郁微": 23 ,"卡米热尼.库尔班": 24 ,"张宸瀚": 25 ,"张伯远": 26 ,"唐棠": 27 ,"邱可": 28 ,"易思": 29 ,"霍浩宇": 30 ,"金恩斌": 31 ,"韩鸣昊": 32 ,"孙中航": 33 ,"王毅杰": 34 ,"德吉央宗": 35 ,"王薪潼": 36 ,"唐堂": 37 ,"李世哲": 38 ,"丁启华": 39 ,"项泉杰": 40 ,"龙雨菲": 41 ,"王周悦": 42 ,"刘晨希": 43 ,"韩志晓": 44 ,"孙兆阳": 45 ,"徐子昂": 46 ,"孙浩": 47 ,"石雅郗": 48 ,"陈睿宁": 49 ,"管意": 50 ,"罗泽宇": 51}
# lis =["郑雨楠","胡博兮","夏逸雨","汪锴","赵皓然","杨帅聪","刘孟乔","周书瑗","李寿媛","陈志辉","杨飒","黄凯威","蓝庭旭","杨致远","黄博闻","鲍云天","刘佳祺","陈千奕","周陇","厍妍","房任博","谭初石","黄郁微","卡米热尼.库尔班","张宸瀚","张伯远","唐棠","邱可","易思","霍浩宇","金恩斌","韩鸣昊","孙中航","王毅杰","德吉央宗","王薪潼","唐堂","李世哲","丁启华","项泉杰","龙雨菲","王周悦","刘晨希","韩志晓","孙兆阳","徐子昂","孙浩","石雅郗","陈睿宁","管意","罗泽宇"]
# dic_tmp={}
# sum_pos = 0
# while 1:
#     str_in = input()
#     if str_in == 'q':
#         break
#     i = str_in.split(',')
#     dic_tmp.setdefault(lis[int(i[0])],[float(sum_pos),float(i[1])+sum_pos])
#     sum_pos += float(i[1])
# final = {}
# while len(final) < 14:
#     pos = random.uniform(0, sum_pos)
#     for k in dic_tmp.keys():
#         if pos < dic_tmp[k][1] and pos >= dic_tmp[k][0]:
#             final.setdefault(k,dic[k])
#             break
# print(final)